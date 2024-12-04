import os
import datetime
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from tkinter import filedialog
import tkinter as tk


def get_next_versioned_filename(base_name, ext):
    """
    버전 번호가 붙은 파일 이름을 생성합니다.
    """
    v = 1
    while True:
        file_name = f"{base_name}_V{v}{ext}"
        if not os.path.exists(file_name):
            return file_name
        v += 1

def set_header_style(header_row):
    """
    헤더 행의 색상과 글꼴 스타일을 설정합니다.
    """
    for cell in header_row:
        cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        cell.font = Font(bold=True)

def handle_upload(file_path):
    """
    암호화된 엑셀 파일을 읽어 데이터를 처리하고 새 엑셀 파일을 생성합니다.
    """
    nowDate = datetime.datetime.now()


    # 파일 이름과 경로 처리
    base, ext = os.path.splitext(file_path)
    decrypted_file_path = f"{base}{ext}"


    try:
        df = pd.read_excel(decrypted_file_path, engine='openpyxl')
    except Exception as e:
        pass
        return

    # 새로운 엑셀 파일 생성
    workbook = Workbook()
    sheet = workbook.active

    headers = ['주문번호', '보내는사람', '받는사람', '전화번호1', '전화번호2', '우편번호', '주소', '품목명', '옵션', '수량', '배송메시지']
    sheet.append(headers)
    set_header_style(sheet[1])

    for _, row in df.iterrows():
        try:
            order_num = "'" + str(row[0])
            buyer = row[10]
            customer = str(row[12]).replace("님", "")
            product = row[18]
            option = row[21]
            amount = int(row[23])
            b_tel = row[43]
            c_tel = row[44]
            address = str(row[45]).strip('"')
            # buyer_phone = row[48]
            message = row[50]

            if buyer == customer:
                buyer = "고객사"
            else:
                buyer = "고객사(" + buyer + ")"
            data = [order_num, buyer, customer + "님", b_tel, c_tel, '', address, product, option, amount, message]
            sheet.append(data)
        except Exception as e:
            print(f"데이터 처리 중 오류 발생: {e}")

    base_excel_file_name = nowDate.strftime("%Y%m%d") + "_스마트스토어_송장양식"
    excel_file_ext = ".xlsx"
    excel_file_name = get_next_versioned_filename(base_excel_file_name, excel_file_ext)

    workbook.save(excel_file_name)
    print(f"엑셀 파일 '{excel_file_name}'이 생성되었습니다.")



def main():

    root = tk.Tk()
    root.withdraw()  # tkinter 창을 숨깁니다.

    file_path = filedialog.askopenfilename(
        title="엑셀 파일 선택",
        filetypes=[("Excel files", "*.xlsx")]  # 엑셀 파일만 선택 가능
    )

    if file_path:
        handle_upload(file_path)
        print("파일 업로드 및 처리가 완료되었습니다.")

if __name__ == "__main__":
    main()
